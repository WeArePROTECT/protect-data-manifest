#!/usr/bin/env python3
"""
PROTECT Data Manifest — Tier-1 enriched crawler.

Reads crawler/collections.yaml (the curated list of data products) and, for each collection,
generates a machine-readable dataset.yaml under manifest/collections/<id>/.

Guarantees:
  * NEVER writes into source data directories (data stays in place).
  * NEVER edits CARD.md (human-owned).
  * Machine facts only: paths, sizes, formats, column names + inferred types, row counts,
    access (from permissions), machine-verified freshness, and a per-subdirectory digest.
  * Column VALUES are never persisted — only headers + inferred types (sensitivity-safe).

Per-collection options (in collections.yaml):
  team, domain            grouping metadata (who owns it / what kind of data)
  resource_globs: [..]    limit which files become resources
  scan_depth: N           how deep to descend for schema extraction (default 1 = top level)
  max_resources: N        cap on described files (default 80)
  deep_size: false        skip the full recursive size walk for very large trees

Deps: stdlib + PyYAML, openpyxl (xlsx), pyarrow (parquet). pandas is intentionally avoided.
"""
import os, sys, csv, re, stat, glob, argparse, datetime, grp
import yaml

csv.field_size_limit(min(sys.maxsize, 2**31 - 1))

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
COLLECTIONS_DIR = os.path.join(REPO, "manifest", "collections")
CONFIG = os.path.join(REPO, "crawler", "collections.yaml")

DATA_EXTS = {".csv", ".tsv", ".xlsx", ".parquet"}
SAMPLE_ROWS = 200
WALK_FILE_CAP = 20000
ROWCOUNT_MAX_BYTES = 50 * 1024 * 1024
MAX_SUBDIRS = 80
KEY_RE = re.compile(r"(^id$|_id$|isolate|sample|patient|subject|asma)", re.I)
# Archive/holding dir names that should NOT be picked as the "newest" subdir — their directory
# mtime bumps whenever an old snapshot is moved into them, which would mask the real current one.
ARCHIVE_DIR_NAMES = {"previous_exports", "previous", "archive", "archived", "old", "backup",
                     "_archive", "prev", "history"}


def iso(ts):
    return datetime.datetime.utcfromtimestamp(ts).strftime("%Y-%m-%dT%H:%M:%SZ")


def now_iso():
    return datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")


def derive_access(path):
    st = os.stat(path)
    world_read = bool(st.st_mode & stat.S_IROTH)
    try:
        group = grp.getgrgid(st.st_gid).gr_name
    except Exception:
        group = str(st.st_gid)
    if world_read:
        return {"status": "open", "readable_by": ["all"], "request_to": None}
    return {"status": "gated", "readable_by": [f"group:{group}"],
            "request_to": f"a member of group '{group}'"}


def infer_type(values):
    if not values:
        return "unknown"

    def ok(fn):
        return all(fn(v) for v in values)

    def is_int(v):
        try:
            int(v); return True
        except Exception:
            return False

    def is_float(v):
        try:
            float(v); return True
        except Exception:
            return False

    def is_date(v):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y", "%Y-%m-%dT%H:%M:%S"):
            try:
                datetime.datetime.strptime(v.strip(), fmt); return True
            except Exception:
                pass
        return False

    def is_bool(v):
        return v.strip().lower() in ("true", "false", "yes", "no")

    if ok(is_int):
        return "integer"
    if ok(is_float):
        return "number"
    if ok(is_date):
        return "date"
    if ok(is_bool):
        return "boolean"
    return "string"


def count_lines(path):
    c = 0
    with open(path, "rb") as f:
        for _ in f:
            c += 1
    return c


def candidate_keys(fields):
    return [f["name"] for f in fields if f["name"] and KEY_RE.search(f["name"])]


def schema_from_delimited(path, delim, size_bytes):
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f, delimiter=delim)
        try:
            header = next(reader)
        except StopIteration:
            return {"fields": [], "candidate_keys": []}, 0
        samples = [[] for _ in header]
        for i, row in enumerate(reader):
            if i >= SAMPLE_ROWS:
                break
            for j, val in enumerate(row[:len(header)]):
                if val != "":
                    samples[j].append(val)
    fields = [{"name": (h or "").strip(), "type": infer_type(samples[j])}
              for j, h in enumerate(header)]
    rows = (count_lines(path) - 1) if size_bytes < ROWCOUNT_MAX_BYTES else None
    return {"fields": fields, "candidate_keys": candidate_keys(fields)}, rows


def resources_from_xlsx(path, size_bytes, rel, mtime):
    import openpyxl
    out = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            try:
                header = list(next(it))
            except StopIteration:
                header = []
            header = [str(h) if h is not None else "" for h in header]
            samples = [[] for _ in header]
            for i, row in enumerate(it):
                if i >= SAMPLE_ROWS:
                    break
                for j, val in enumerate(row[:len(header)]):
                    if val not in (None, ""):
                        samples[j].append(str(val))
            fields = [{"name": header[j].strip(), "type": infer_type(samples[j])}
                      for j in range(len(header))]
            out.append({"path": rel, "format": "xlsx", "size_bytes": size_bytes, "mtime": mtime,
                        "rows": ws.max_row, "sheet": ws.title,
                        "schema": {"fields": fields, "candidate_keys": candidate_keys(fields)}})
    finally:
        wb.close()
    return out


def schema_from_parquet(path):
    import pyarrow.parquet as pq
    pf = pq.ParquetFile(path)
    sch = pf.schema_arrow
    fields = [{"name": n, "type": str(sch.field(n).type)} for n in sch.names]
    return {"fields": fields, "candidate_keys": candidate_keys(fields)}, pf.metadata.num_rows


def collect_data_files(root, scan_depth, globs, exclude_dirs=None):
    """Data files within scan_depth levels of root (default 1 = top level only).

    exclude_dirs: directory names to prune from the walk (e.g. huge per-isolate trees),
    so high-value analysis files aren't crowded out of the max_resources cap.
    """
    if globs:
        files = []
        for g in globs:
            files.extend(glob.glob(os.path.join(root, g)))
        return sorted(set(files))
    exclude = set(exclude_dirs or [])
    base = root.rstrip(os.sep).count(os.sep)
    files = []
    for dirpath, dirs, fnames in os.walk(root):
        dirs[:] = [d for d in dirs if d not in exclude]
        depth = dirpath.rstrip(os.sep).count(os.sep) - base
        if depth >= scan_depth:
            dirs[:] = []  # stop descending
        for fn in fnames:
            if os.path.splitext(fn)[1].lower() in DATA_EXTS:
                files.append(os.path.join(dirpath, fn))
    return sorted(set(files))


def subdir_digests(root):
    """Shallow (one-level) digest of each immediate subdirectory — fast and safe on huge trees."""
    out = []
    try:
        entries = sorted(os.scandir(root), key=lambda e: e.name)
    except OSError:
        return out
    for entry in entries:
        try:
            if not entry.is_dir(follow_symlinks=False):
                continue
        except OSError:
            continue
        files = subs = total = 0
        exts = {}
        try:
            for e in os.scandir(entry.path):
                try:
                    if e.is_dir(follow_symlinks=False):
                        subs += 1
                    elif e.is_file(follow_symlinks=False):
                        files += 1
                        try:
                            total += e.stat(follow_symlinks=False).st_size
                        except OSError:
                            pass
                        ext = os.path.splitext(e.name)[1].lower() or "(noext)"
                        exts[ext] = exts.get(ext, 0) + 1
                except OSError:
                    continue
        except OSError:
            pass
        try:
            sub_mtime = iso(entry.stat(follow_symlinks=False).st_mtime)
        except OSError:
            sub_mtime = None
        out.append({"name": entry.name, "mtime": sub_mtime,
                    "immediate_files": files, "immediate_subdirs": subs,
                    "immediate_bytes": total,
                    "file_types": dict(sorted(exts.items(), key=lambda kv: -kv[1])[:6])})
        if len(out) >= MAX_SUBDIRS:
            break
    return out


def find_owner_manifest(root):
    for base in (root, os.path.dirname(root)):
        for fn in ("MANIFEST.md", "README.md"):
            p = os.path.join(base, fn)
            if os.path.isfile(p):
                claimed = None
                try:
                    with open(p, encoding="utf-8", errors="replace") as f:
                        head = f.read(2000)
                    m = re.search(r"last[_ ]updated\s*[:=]\s*([0-9]{4}-[0-9]{2}-[0-9]{2})", head, re.I)
                    if m:
                        claimed = m.group(1)
                except OSError:
                    pass
                return {"path": p, "last_updated_claimed": claimed}
    return None


def walk_size(root, do_deep):
    total = count = 0
    latest = 0.0
    truncated = not do_deep
    if not do_deep:
        try:
            for name in os.listdir(root):
                try:
                    st = os.lstat(os.path.join(root, name))
                except OSError:
                    continue
                if stat.S_ISREG(st.st_mode):
                    total += st.st_size
                    count += 1
                latest = max(latest, st.st_mtime)
        except OSError:
            pass
        return total, count, latest, truncated
    for dirpath, _dirs, files in os.walk(root):
        for fn in files:
            try:
                st = os.lstat(os.path.join(dirpath, fn))
            except OSError:
                continue
            total += st.st_size
            count += 1
            latest = max(latest, st.st_mtime)
            if count >= WALK_FILE_CAP:
                return total, count, latest, True
    return total, count, latest, truncated


def crawl_collection(c):
    root = c["collection_root"]
    out = {"id": c["id"], "title": c.get("title", c["id"]),
           "collection_root": root,
           "team": c.get("team"), "domain": c.get("domain"),
           "owner": c.get("owner", {}), "tags": c.get("tags") or [], "notes": []}
    if not os.path.exists(root):
        out["access"] = {"status": "unknown", "readable_by": [], "request_to": None}
        out["notes"].append(f"collection_root does not exist: {root}")
        return out
    out["access"] = derive_access(root)
    if c.get("request_to"):  # human-provided steward overrides the generic group hint
        out["access"]["request_to"] = c["request_to"]

    scan_depth = c.get("scan_depth", 1)
    max_resources = c.get("max_resources", 80)
    # pick_latest_subdir: for dated-run dirs, describe the NEWEST run (by mtime) while still
    # digesting every run below — so the catalog auto-follows the current run, no date in config.
    resource_root = root
    if c.get("pick_latest_subdir"):
        try:
            subs = [os.path.join(root, n) for n in os.listdir(root)
                    if os.path.isdir(os.path.join(root, n))]
        except OSError:
            subs = []
        if subs:
            resource_root = max(subs, key=lambda p: os.stat(p).st_mtime)
            out["notes"].append(f"resources taken from latest run: {os.path.relpath(resource_root, root)}")
    files = collect_data_files(resource_root, scan_depth, c.get("resource_globs"), c.get("exclude_dirs"))
    if len(files) > max_resources:
        out["notes"].append(f"resource list truncated to {max_resources} of {len(files)} data files")
        files = files[:max_resources]

    resources = []
    for p in files:
        ext = os.path.splitext(p)[1].lower()
        rel = os.path.relpath(p, root)
        try:
            size_b = os.path.getsize(p)
            mtime = iso(os.path.getmtime(p))   # real per-file modified date (not parsed from filename)
            if ext == ".csv":
                schema, rows = schema_from_delimited(p, ",", size_b)
                resources.append({"path": rel, "format": "csv", "size_bytes": size_b, "mtime": mtime,
                                  "rows": rows, "sheet": None, "schema": schema})
            elif ext == ".tsv":
                schema, rows = schema_from_delimited(p, "\t", size_b)
                resources.append({"path": rel, "format": "tsv", "size_bytes": size_b, "mtime": mtime,
                                  "rows": rows, "sheet": None, "schema": schema})
            elif ext == ".xlsx":
                resources.extend(resources_from_xlsx(p, size_b, rel, mtime))
            elif ext == ".parquet":
                schema, rows = schema_from_parquet(p)
                resources.append({"path": rel, "format": "parquet", "size_bytes": size_b, "mtime": mtime,
                                  "rows": rows, "sheet": None, "schema": schema})
        except Exception as e:
            out["notes"].append(f"schema extraction failed for {rel}: {e.__class__.__name__}: {e}")
    out["resources"] = resources
    # Name the most-recently-modified resource so "which file is current/newest" is a STATED
    # machine fact (refreshed every crawl), not something an agent must infer from messy filenames.
    dated = [r for r in resources if r.get("mtime")]
    out["latest_resource"] = max(dated, key=lambda r: r["mtime"])["path"] if dated else None

    total, count, latest, trunc = walk_size(root, c.get("deep_size", True))
    try:
        latest = max(latest, os.stat(root).st_mtime)
    except OSError:
        pass
    tree_source_iso = iso(latest) if latest else None
    if c.get("resource_globs"):
        # Glob-scoped collection: the product IS the matched file(s), even though collection_root is a
        # broader shared dir (e.g. the sample roster is one CSV living inside all of Zengler_Lab).
        # Derive size + source mtime from the matched resources — and skip the subdir digest — so
        # unrelated sibling activity in that shared root can't trigger false "changed"/"stale" flags.
        by_path = {r["path"]: r.get("size_bytes", 0) for r in resources}
        res_mtimes = [r["mtime"] for r in resources if r.get("mtime")]
        out["size"] = {"total_bytes": sum(by_path.values()), "file_count": len(by_path),
                       "truncated": False}
        out["subdirectories"] = []
        out["latest_subdir"] = None
        source_iso = max(res_mtimes) if res_mtimes else tree_source_iso
    else:
        out["size"] = {"total_bytes": total, "file_count": count, "truncated": trunc}
        out["subdirectories"] = subdir_digests(root)
        # Newest immediate subdirectory by mtime — for dated-export/dated-run collections this names the
        # current snapshot directly. Archive folders (e.g. previous_exports) are excluded so they don't
        # masquerade as "newest" just because something was recently moved into them. Null if none apply.
        dated_subs = [s for s in out["subdirectories"]
                      if s.get("mtime") and s["name"].lower() not in ARCHIVE_DIR_NAMES]
        out["latest_subdir"] = max(dated_subs, key=lambda s: s["mtime"])["name"] if dated_subs else None
        source_iso = tree_source_iso
    out["freshness"] = {"last_crawled": now_iso(),
                        "source_mtime_latest": source_iso,
                        "cadence_observed": None}
    out["owner_manifest"] = find_owner_manifest(root)
    return out


def write_dataset_yaml(out):
    d = os.path.join(COLLECTIONS_DIR, out["id"])
    os.makedirs(d, exist_ok=True)
    path = os.path.join(d, "dataset.yaml")
    with open(path, "w", encoding="utf-8") as f:
        f.write("# AUTO-GENERATED by crawler/crawl.py — do not hand-edit (regenerated each run).\n")
        f.write("# Human narrative lives in the sibling CARD.md, which the crawler never touches.\n")
        yaml.safe_dump(out, f, sort_keys=False, allow_unicode=True, default_flow_style=False, width=100)
    return path


def main():
    ap = argparse.ArgumentParser(description="PROTECT manifest Tier-1 crawler")
    ap.add_argument("--only", help="crawl only this collection id")
    ap.add_argument("--config", default=CONFIG)
    args = ap.parse_args()

    with open(args.config) as f:
        cfg = yaml.safe_load(f)

    n = 0
    for c in cfg.get("collections", []):
        if args.only and c["id"] != args.only:
            continue
        out = crawl_collection(c)
        path = write_dataset_yaml(out)
        nfields = sum(len(r["schema"]["fields"]) for r in out.get("resources", []))
        nsub = len(out.get("subdirectories", []))
        print(f"[ok] {out['id']}: {len(out.get('resources', []))} resources, {nfields} fields, "
              f"{nsub} subdirs, access={out['access']['status']} -> {os.path.relpath(path, REPO)}")
        for note in out.get("notes", []):
            print(f"     note: {note}")
        n += 1
    print(f"\nCrawled {n} collection(s).")


if __name__ == "__main__":
    main()
